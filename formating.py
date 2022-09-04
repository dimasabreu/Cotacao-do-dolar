from openpyxl.styles import Font, Alignment
import openpyxl


def treatment():
    wb = openpyxl.load_workbook('Countries list.xlsx')
    ws = wb.active
    ws.insert_rows(idx=1, amount=1)
    ws.merge_cells('A1:C1')
    cell = ws.cell(row=1, column=1)
    cell.value = 'Countries List'
    cell.alignment = Alignment(horizontal="center")
    cell2 = ws['A1']
    cell2.font = Font(size=16, bold=True, color='4F4F4F')
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ft = Font(size=12, bold=True, color="808080")
    a2 = ws['A2']
    b2 = ws['B2']
    c2 = ws['C2']
    a2.font = ft
    b2.font = ft
    c2.font = ft
    i=3
    wb.save('Countries list.xlsx')
    return


if __name__ == '__main__':
    treatment()
